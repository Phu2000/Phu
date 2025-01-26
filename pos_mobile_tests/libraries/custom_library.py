import openpyxl

# desired_caps = {
#     "platformName": "Android",
#     "deviceName": "R58T21APQLW",
#     "appPackage": "th.co.gosoft.posonline",
#     "appActivity": "th.co.gosoft.posonline.MainActivity",
#     "automationName": "UiAutomator2"
# }

#Read column fron excel
def ReadColumnFromExcel(file_path, sheet_name, column_name):
    """
    Read data from a specific column in an Excel file.

    :param file_path: Path to the Excel file.
    :param sheet_name: Name of the sheet to read from.
    :param column_name: The header of the column to read data from.
    :return: A list of values from the specified column.
    """
    try:
        # Load the workbook and select the sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Find the column index for the given column name
        column_index = None
        for col in sheet.iter_cols(1, sheet.max_column):
            if col[0].value == column_name:
                column_index = col[0].column
                break

        if column_index is None:
            raise ValueError(f"Column '{column_name}' not found in the sheet.")

        # Read all values from the column using iter_rows instead of iter_cols
        data = [sheet.cell(row=row, column=column_index).value for row in range(2, sheet.max_row + 1)]

        # Return the filtered list, removing None values
        return [item for item in data if item is not None]

    except Exception as e:
        raise RuntimeError(f"Error reading Excel file: {str(e)}")

def tap_with_position(x, y):
        from appium.webdriver.common.touch_action import TouchAction
        from robot.libraries.BuiltIn import BuiltIn

        # Get Appium driver instance
        driver = BuiltIn().get_library_instance("AppiumLibrary")._current_application()
        touch_action = TouchAction(driver)

        # Perform tap action
        touch_action.tap(x=x, y=y).perform()
